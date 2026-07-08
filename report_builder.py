# -*- coding: utf-8 -*-
"""
Генератор Excel-отчёта по аварийным ситуациям на сети.

build_report(raw_records, output_path) собирает 9 листов (см. PRD, раздел 4):
Дашборд, По городам, По зонам ответственности, Топ проблемных объектов,
По характеру повреждения, Динамика по дням, Текущие аварии, По объектам,
Сырые данные — с оформлением, условным форматированием и диаграммами.

build_alert_text(records) формирует текст telegram-сводки по авариям
с простоем более CRITICAL_THRESHOLD_MINUTES минут (см. PRD, раздел 5).
"""

import json
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from parser import deduplicate, format_minutes

CRITICAL_THRESHOLD_MINUTES = 480  # 8 часов

# --- Оформление -------------------------------------------------------

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
BOLD_DATA_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True)
KPI_LABEL_FONT = Font(name="Arial", size=10, bold=True)
KPI_VALUE_FONT = Font(name="Arial", size=18, bold=True, color="1F4E78")

RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
RED_FONT = Font(name="Arial", size=10, color="9C0006")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")

THIN_SIDE = Side(style="thin", color="B7B7B7")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center")

TAB_COLORS = {
    "Дашборд": "1F4E78",
    "По городам": "2E75B6",
    "По зонам ответственности": "2E75B6",
    "Топ проблемных объектов": "C00000",
    "По характеру повреждения": "2E75B6",
    "Динамика по дням": "2E75B6",
    "Текущие аварии": "ED7D31",
    "По объектам": "548235",
    "Сырые данные": "808080",
}


def _fmt_date(dt):
    return dt.strftime("%d.%m.%Y %H:%M:%S") if dt else ""


def _write_header_row(ws, row, headers):
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_data_row(ws, row, values, bold=False):
    font = BOLD_DATA_FONT if bold else DATA_FONT
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.border = THIN_BORDER


def _autosize(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _apply_duration_conditional_formatting(ws, col_letter, first_row, last_row):
    if last_row < first_row:
        return
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=[str(CRITICAL_THRESHOLD_MINUTES)],
            fill=RED_FILL,
            font=RED_FONT,
        ),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="lessThan",
            formula=[str(CRITICAL_THRESHOLD_MINUTES)],
            fill=GREEN_FILL,
        ),
    )


def _set_tab_colors(wb):
    for title, color in TAB_COLORS.items():
        if title in wb.sheetnames:
            wb[title].sheet_properties.tabColor = color


# --- Агрегация ----------------------------------------------------------

def _aggregate_by(records, key_func):
    """Группирует записи по key_func(record) и считает базовую статистику."""
    groups = {}
    for rec in records:
        key = key_func(rec) or "—"
        g = groups.setdefault(
            key, {"key": key, "count": 0, "count_gt8h": 0, "total_minutes": 0}
        )
        g["count"] += 1
        minutes = rec.get("duration_minutes") or 0
        g["total_minutes"] += minutes
        if minutes >= CRITICAL_THRESHOLD_MINUTES:
            g["count_gt8h"] += 1
    result = list(groups.values())
    for g in result:
        g["avg_minutes"] = g["total_minutes"] / g["count"] if g["count"] else 0
    result.sort(key=lambda g: g["total_minutes"], reverse=True)
    return result


def _aggregate_nodes(records):
    groups = {}
    for rec in records:
        key = rec.get("node") or "—"
        g = groups.setdefault(
            key,
            {
                "node": key,
                "city": rec.get("city", ""),
                "zone": rec.get("zone", ""),
                "count": 0,
                "total_minutes": 0,
                "max_minutes": 0,
            },
        )
        minutes = rec.get("duration_minutes") or 0
        g["count"] += 1
        g["total_minutes"] += minutes
        g["max_minutes"] = max(g["max_minutes"], minutes)
    result = list(groups.values())
    result.sort(key=lambda g: g["total_minutes"], reverse=True)
    return result


def _incident_date(rec):
    if rec.get("created_at"):
        return rec["created_at"].date()
    if rec.get("message_date"):
        return rec["message_date"].date()
    return None


def get_escalation_records(records, threshold=CRITICAL_THRESHOLD_MINUTES):
    """Дедуплицированные аварии с простоем >= threshold минут, по убыванию простоя."""
    deduped = deduplicate(records)
    escalated = [r for r in deduped if (r.get("duration_minutes") or 0) >= threshold]
    escalated.sort(key=lambda r: r.get("duration_minutes") or 0, reverse=True)
    return escalated


# --- Листы ----------------------------------------------------------------

def _write_dashboard(wb, deduped):
    ws = wb.create_sheet("Дашборд")
    ws["B2"] = "Отчёт по аварийным ситуациям на сети"
    ws["B2"].font = TITLE_FONT

    total = len(deduped)
    total_minutes = sum(r.get("duration_minutes") or 0 for r in deduped)
    count_gt8h = sum(1 for r in deduped if (r.get("duration_minutes") or 0) >= CRITICAL_THRESHOLD_MINUTES)
    avg_minutes = total_minutes / total if total else 0

    by_city = _aggregate_by(deduped, lambda r: r.get("city"))
    problem_city = by_city[0]["key"] if by_city else "—"

    by_node_count = sorted(
        _aggregate_nodes(deduped), key=lambda g: g["count"], reverse=True
    )
    frequent_node = by_node_count[0]["node"] if by_node_count else "—"

    kpis = [
        ("Всего аварий", total),
        ("Простой >8ч", count_gt8h),
        ("Общий простой", format_minutes(total_minutes)),
        ("Средний простой", format_minutes(round(avg_minutes))),
        ("Проблемный город", problem_city),
        ("Частый объект", frequent_node),
    ]

    row = 4
    for label, value in kpis:
        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.font = KPI_LABEL_FONT
        value_cell = ws.cell(row=row, column=4, value=value)
        value_cell.font = KPI_VALUE_FONT
        row += 2

    _autosize(ws, [4, 24, 4, 30])
    return ws


def _write_city_sheet(wb, deduped):
    ws = wb.create_sheet("По городам")
    headers = ["Город", "Кол-во аварий", "Простой >8ч", "Общий простой", "Простой (мин.)", "Средний простой"]
    _write_header_row(ws, 1, headers)

    groups = _aggregate_by(deduped, lambda r: r.get("city"))
    row = 2
    for g in groups:
        _write_data_row(ws, row, [
            g["key"], g["count"], g["count_gt8h"],
            format_minutes(g["total_minutes"]), g["total_minutes"],
            format_minutes(round(g["avg_minutes"])),
        ])
        row += 1

    total_row = row
    _write_data_row(ws, total_row, [
        "ИТОГО", sum(g["count"] for g in groups), sum(g["count_gt8h"] for g in groups),
        format_minutes(sum(g["total_minutes"] for g in groups)),
        sum(g["total_minutes"] for g in groups), "",
    ], bold=True)

    _apply_duration_conditional_formatting(ws, "E", 2, total_row - 1)
    _autosize(ws, [22, 14, 14, 20, 15, 18])

    if len(groups) > 0:
        chart = BarChart()
        chart.title = "Количество аварий по городам"
        chart.y_axis.title = "Кол-во аварий"
        chart.x_axis.title = "Город"
        data = Reference(ws, min_col=2, min_row=1, max_row=total_row - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=total_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 20, 11
        ws.add_chart(chart, "H2")

    return ws


def _write_zone_sheet(wb, deduped):
    ws = wb.create_sheet("По зонам ответственности")
    headers = ["Зона ответственности", "Кол-во аварий", "Простой >8ч", "Общий простой", "Простой (мин.)", "Средний простой"]
    _write_header_row(ws, 1, headers)

    groups = _aggregate_by(deduped, lambda r: r.get("zone"))
    row = 2
    for g in groups:
        _write_data_row(ws, row, [
            g["key"], g["count"], g["count_gt8h"],
            format_minutes(g["total_minutes"]), g["total_minutes"],
            format_minutes(round(g["avg_minutes"])),
        ])
        row += 1

    total_row = row
    _write_data_row(ws, total_row, [
        "ИТОГО", sum(g["count"] for g in groups), sum(g["count_gt8h"] for g in groups),
        format_minutes(sum(g["total_minutes"] for g in groups)),
        sum(g["total_minutes"] for g in groups), "",
    ], bold=True)

    _apply_duration_conditional_formatting(ws, "E", 2, total_row - 1)
    _autosize(ws, [32, 14, 14, 20, 15, 18])

    if len(groups) > 0:
        chart = PieChart()
        chart.title = "Распределение аварий по зонам ответственности"
        data = Reference(ws, min_col=2, min_row=1, max_row=total_row - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=total_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 20, 11
        ws.add_chart(chart, "H2")

    return ws


def _write_top_nodes_sheet(wb, deduped):
    ws = wb.create_sheet("Топ проблемных объектов")
    headers = [
        "Узел сети", "Город", "Зона ответственности", "Кол-во аварий",
        "Общий простой", "Простой (мин.)", "Макс. простой одной аварии",
    ]
    _write_header_row(ws, 1, headers)

    groups = _aggregate_nodes(deduped)[:30]
    row = 2
    for g in groups:
        _write_data_row(ws, row, [
            g["node"], g["city"], g["zone"], g["count"],
            format_minutes(g["total_minutes"]), g["total_minutes"],
            format_minutes(g["max_minutes"]),
        ])
        row += 1

    _apply_duration_conditional_formatting(ws, "F", 2, row - 1)
    _autosize(ws, [30, 18, 26, 14, 20, 15, 26])
    return ws


def _write_damage_type_sheet(wb, deduped):
    ws = wb.create_sheet("По характеру повреждения")
    headers = ["Характер повреждения", "Кол-во аварий", "Общий простой", "Простой (мин.)", "Средний простой"]
    _write_header_row(ws, 1, headers)

    groups = _aggregate_by(deduped, lambda r: r.get("damage_type"))
    row = 2
    for g in groups:
        _write_data_row(ws, row, [
            g["key"], g["count"], format_minutes(g["total_minutes"]),
            g["total_minutes"], format_minutes(round(g["avg_minutes"])),
        ])
        row += 1

    _apply_duration_conditional_formatting(ws, "D", 2, row - 1)
    _autosize(ws, [24, 14, 20, 15, 18])
    return ws


def _write_daily_sheet(wb, deduped):
    ws = wb.create_sheet("Динамика по дням")
    headers = ["Дата", "Кол-во аварий", "Простой >8ч", "Суммарный простой", "Простой (мин.)"]
    _write_header_row(ws, 1, headers)

    groups = _aggregate_by(deduped, _incident_date)
    groups.sort(key=lambda g: (g["key"] == "—", g["key"]))

    row = 2
    for g in groups:
        date_val = g["key"].strftime("%d.%m.%Y") if hasattr(g["key"], "strftime") else g["key"]
        _write_data_row(ws, row, [
            date_val, g["count"], g["count_gt8h"],
            format_minutes(g["total_minutes"]), g["total_minutes"],
        ])
        row += 1

    total_row = row
    _apply_duration_conditional_formatting(ws, "E", 2, total_row - 1)
    _autosize(ws, [16, 14, 14, 20, 15])

    if row > 2:
        chart = BarChart()
        chart.title = "Аварии по дням"
        chart.y_axis.title = "Кол-во аварий"
        chart.x_axis.title = "Дата"
        data = Reference(ws, min_col=2, min_row=1, max_row=total_row - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=total_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 20, 11
        ws.add_chart(chart, "H2")

    return ws


def _write_current_incidents_sheet(wb, deduped, now):
    ws = wb.create_sheet("Текущие аварии")
    headers = [
        "Город", "Узел сети", "Номер ПБ", "Время возникновения",
        "Простой", "Простой (мин.)", "Зона ответственности", "Статус",
    ]
    _write_header_row(ws, 1, headers)

    cutoff = now - timedelta(hours=2)
    current = [
        r for r in deduped
        if r.get("message_date") and r["message_date"] >= cutoff
    ]
    current.sort(key=lambda r: r.get("duration_minutes") or 0, reverse=True)

    row = 2
    for r in current:
        minutes = r.get("duration_minutes") or 0
        status = "Простой >8ч" if minutes >= CRITICAL_THRESHOLD_MINUTES else "В работе"
        _write_data_row(ws, row, [
            r.get("city", ""), r.get("node", ""), r.get("pb_number", ""),
            _fmt_date(r.get("created_at")) or r.get("created_at_raw", ""),
            r.get("duration_raw", ""), minutes, r.get("zone", ""), status,
        ])
        row += 1

    _apply_duration_conditional_formatting(ws, "F", 2, row - 1)
    _autosize(ws, [20, 30, 20, 20, 20, 15, 30, 16])
    return ws


def _write_by_object_sheet(wb, deduped):
    ws = wb.create_sheet("По объектам")
    headers = [
        "Город", "Узел сети", "Зона ответственности", "Номер ПБ",
        "Время возникновения", "Характер повреждения", "Простой", "Простой (мин.)",
    ]
    _write_header_row(ws, 1, headers)

    ordered = sorted(deduped, key=lambda r: (r.get("city") or "", r.get("node") or ""))
    row = 2
    for r in ordered:
        _write_data_row(ws, row, [
            r.get("city", ""), r.get("node", ""), r.get("zone", ""), r.get("pb_number", ""),
            _fmt_date(r.get("created_at")) or r.get("created_at_raw", ""),
            r.get("damage_type", ""), r.get("duration_raw", ""), r.get("duration_minutes") or 0,
        ])
        row += 1

    _apply_duration_conditional_formatting(ws, "H", 2, row - 1)
    _autosize(ws, [20, 30, 26, 20, 20, 18, 20, 15])
    return ws


def _write_raw_data_sheet(wb, raw_records):
    ws = wb.create_sheet("Сырые данные")
    headers = [
        "Подчат", "Город", "Район", "Узел сети", "Номер ПБ", "Время возникновения",
        "Характер повреждения", "Источник аварии", "Описание", "Длительность", "Длительность (мин.)",
    ]
    _write_header_row(ws, 1, headers)

    ordered = sorted(
        raw_records,
        key=lambda r: (r.get("subchat") or "", r.get("pb_number") or "", r.get("message_date") or datetime.min),
    )
    row = 2
    for r in ordered:
        _write_data_row(ws, row, [
            r.get("subchat", ""), r.get("city", ""), r.get("district", ""), r.get("node", ""),
            r.get("pb_number", ""), _fmt_date(r.get("created_at")) or r.get("created_at_raw", ""),
            r.get("damage_type", ""), r.get("source", ""), r.get("description", ""),
            r.get("duration_raw", ""), r.get("duration_minutes") or 0,
        ])
        row += 1

    _apply_duration_conditional_formatting(ws, "K", 2, row - 1)
    _autosize(ws, [24, 18, 18, 30, 20, 20, 18, 26, 34, 20, 18])
    return ws


# --- Публичное API ----------------------------------------------------------

def build_report(raw_records, output_path, now=None):
    """Строит .xlsx-отчёт из сырых (недедуплицированных) записей об авариях.

    raw_records: список словарей в формате parser.parse_message().
    output_path: путь к выходному .xlsx файлу.
    now: точка отсчёта для листа "Текущие аварии" (по умолчанию — текущее время).
    """
    now = now or datetime.now()
    deduped = deduplicate(raw_records)

    wb = Workbook()
    wb.remove(wb.active)  # убираем дефолтный пустой лист, добавляем листы по порядку из ТЗ

    _write_dashboard(wb, deduped)
    _write_city_sheet(wb, deduped)
    _write_zone_sheet(wb, deduped)
    _write_top_nodes_sheet(wb, deduped)
    _write_damage_type_sheet(wb, deduped)
    _write_daily_sheet(wb, deduped)
    _write_current_incidents_sheet(wb, deduped, now)
    _write_by_object_sheet(wb, deduped)
    _write_raw_data_sheet(wb, raw_records)

    _set_tab_colors(wb)
    wb.save(output_path)
    return output_path


def build_alert_text(raw_records, threshold=CRITICAL_THRESHOLD_MINUTES):
    """Формирует текст telegram-сводки по авариям с простоем >= threshold минут.

    Возвращает None, если таких аварий нет (сообщение отправлять не нужно).
    """
    escalated = get_escalation_records(raw_records, threshold)
    if not escalated:
        return None

    lines = [
        "🚨 СВОДКА: ПРОСТОЙ БОЛЕЕ 8 ЧАСОВ",
        f"Всего аварий: {len(escalated)}",
        "━━━━━━━━━━━━━━━━━━━━━",
    ]
    for r in escalated:
        lines.append("")
        lines.append("🔴 ПРОСТОЙ БОЛЕЕ 8 ЧАСОВ")
        lines.append("━━━━━━━━━━━━━━━━━━━━━")
        lines.append(f"📍 Город: {r.get('city', '')}")
        lines.append(f"🏗 Объект: {r.get('node', '')}")
        lines.append(f"⏱ Простой: {r.get('duration_raw', '')}")
        lines.append(f"🏢 Зона ответственности: {r.get('zone', '')}")
        lines.append(f"📋 Номер ПБ: {r.get('pb_number', '')}")
        lines.append(f"🕐 Начало: {_fmt_date(r.get('created_at')) or r.get('created_at_raw', '')}")
        lines.append(f"🔧 Характер: {r.get('damage_type', '')}")
        lines.append("━━━━━━━━━━━━━━━━━━━━━")

    return "\n".join(lines)


def _serialize_groups(groups, key_name):
    result = []
    for g in groups:
        key = g["key"]
        key = key.strftime("%d.%m.%Y") if hasattr(key, "strftime") else key
        result.append({
            key_name: key,
            "count": g["count"],
            "count_gt8h": g["count_gt8h"],
            "total_minutes": g["total_minutes"],
            "avg_minutes": round(g["avg_minutes"]),
        })
    return result


def build_dashboard_data(raw_records, now=None):
    """Собирает те же агрегаты, что и Excel-отчёт, в виде JSON-совместимого словаря
    (для веб-дашборда, см. docs/)."""
    now = now or datetime.now()
    deduped = deduplicate(raw_records)

    total = len(deduped)
    total_minutes = sum(r.get("duration_minutes") or 0 for r in deduped)
    count_gt8h = sum(1 for r in deduped if (r.get("duration_minutes") or 0) >= CRITICAL_THRESHOLD_MINUTES)
    avg_minutes = total_minutes / total if total else 0

    by_city = _aggregate_by(deduped, lambda r: r.get("city"))
    by_zone = _aggregate_by(deduped, lambda r: r.get("zone"))
    by_damage = _aggregate_by(deduped, lambda r: r.get("damage_type"))
    daily = _aggregate_by(deduped, _incident_date)
    daily.sort(key=lambda g: (g["key"] == "—", g["key"]))
    top_nodes = _aggregate_nodes(deduped)[:10]

    problem_city = by_city[0]["key"] if by_city else "—"
    by_node_count = sorted(_aggregate_nodes(deduped), key=lambda g: g["count"], reverse=True)
    frequent_node = by_node_count[0]["node"] if by_node_count else "—"

    escalations = get_escalation_records(raw_records, CRITICAL_THRESHOLD_MINUTES)

    return {
        "generated_at": now.isoformat(),
        "kpi": {
            "total_incidents": total,
            "gt8h_count": count_gt8h,
            "total_downtime_minutes": total_minutes,
            "avg_downtime_minutes": round(avg_minutes),
            "problem_city": problem_city,
            "frequent_node": frequent_node,
        },
        "by_city": _serialize_groups(by_city, "city"),
        "by_zone": _serialize_groups(by_zone, "zone"),
        "by_damage_type": _serialize_groups(by_damage, "damage_type"),
        "daily": _serialize_groups(daily, "date"),
        "top_nodes": [
            {
                "node": g["node"], "city": g["city"], "zone": g["zone"],
                "count": g["count"], "total_minutes": g["total_minutes"],
                "max_minutes": g["max_minutes"],
            }
            for g in top_nodes
        ],
        "escalations": [
            {
                "pb_number": r.get("pb_number", ""),
                "city": r.get("city", ""),
                "node": r.get("node", ""),
                "zone": r.get("zone", ""),
                "duration_text": r.get("duration_raw", ""),
                "duration_minutes": r.get("duration_minutes") or 0,
                "started_at": _fmt_date(r.get("created_at")) or r.get("created_at_raw", ""),
                "damage_type": r.get("damage_type", ""),
            }
            for r in escalations
        ],
    }


def export_dashboard_json(raw_records, path, now=None):
    """Пишет build_dashboard_data(...) в файл path в формате JSON (UTF-8)."""
    data = build_dashboard_data(raw_records, now=now)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path

    return "\n".join(lines)
